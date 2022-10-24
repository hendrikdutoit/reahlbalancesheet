from reahl.web.fw import UserInterface, Widget
from reahl.web.bootstrap.page import HTML5Page
from reahl.web.bootstrap.ui import Div, H, P, FieldSet
from reahl.web.bootstrap.navbar import Navbar, ResponsiveLayout
from reahl.web.bootstrap.navs import Nav
from reahl.web.bootstrap.grid import Container
from reahl.web.bootstrap.forms import TextInput, Form, FormLayout, Button
from reahl.web.bootstrap.files import FileUploadInput
from reahl.component.modelinterface import (
    exposed,
    Field,
    Action,
    Event,
    DateField,
    FileField,
)
from reahl.sqlalchemysupport import Session, Base
from reahl.web.plotly import Chart
import plotly.graph_objects as pg
from sqlalchemy import Column, Integer, UnicodeText, Date, LargeBinary, func
from openpyxl import load_workbook
from io import BytesIO


class CustomerPage(HTML5Page):
    def __init__(self, view, bookmarks):
        super().__init__(view)
        self.body.use_layout(Container())

        layout = ResponsiveLayout('md', colour_theme='dark', bg_scheme='primary')
        navbar = Navbar(view, css_id='my_nav').use_layout(layout)
        navbar.layout.set_brand_text('Customers')
        # navbar.layout.add(TextNode(view, 'All your customers'))
        navbar.layout.add(Nav(view).with_bookmarks(bookmarks))

        self.body.add_child(navbar)
        self.curr_rec_id = None
        # self.body.add_child(CustomerBookPanel(view))


class HomePage(CustomerPage):
    def __init__(self, view, main_bookmarks):
        super().__init__(view, main_bookmarks)
        self.body.add_child(CustomerPanel(view))


class AddCustomerPage(CustomerPage):
    def __init__(self, view, main_bookmarks):
        super().__init__(view, main_bookmarks)
        self.body.add_child(CustomerForm(view))


class CustomerForm(Form):
    def __init__(self, view):
        super().__init__(view, 'customer_form')

        inputs = self.add_child(FieldSet(view, legend_text='Add a customer'))
        inputs.use_layout(FormLayout())

        new_customer = Customer()
        inputs.layout.add_input(TextInput(self, new_customer.fields.surname))
        inputs.layout.add_input(TextInput(self, new_customer.fields.name))
        inputs.layout.add_input(TextInput(self, new_customer.fields.dob))

        inputs.layout.add_input(
            FileUploadInput(self, new_customer.fields.uploaded_files), hide_label=True
        )

        inputs.add_child(Button(self, new_customer.events.save))


class CustomerPanel(Div):
    def __init__(self, view):
        super().__init__(view)

        self.add_child(H(view, 1, text='Customers'))

        # self.add_child(CustomerForm(view))

        for customer in Session.query(Customer).all():
            self.add_child(CustomerBox(view, customer))


class CustomerBox(Widget):
    def __init__(self, view, customer):
        super().__init__(view)
        self.add_child(
            P(
                view,
                text='%s %s, %s - %s'
                % (customer.dob, customer.surname, customer.name, customer.bsfilename),
            )
        )


class CustomerUI(UserInterface):
    def assemble(self):
        home = self.define_view('/', title='Show')
        add = self.define_view('/add', title='Add')
        graph = self.define_view('/graph', title='Graph')

        # self.define_transition(Customer.events.save, home, home)
        bookmarks = [v.as_bookmark(self) for v in [home, add, graph]]
        home.set_page(HomePage.factory(bookmarks))
        add.set_page(AddCustomerPage.factory(bookmarks))
        graph.set_page(GraphPage.factory(bookmarks))

        self.define_transition(Customer.events.save, add, graph)
        # self.define_transition(GraphPage.events.back, graph, home)


class Customer(Base):
    __tablename__ = 'customer'

    id = Column(Integer, primary_key=True)
    surname = Column(UnicodeText)
    name = Column(UnicodeText)
    dob = Column(Date)
    bsfilename = Column(UnicodeText)
    bscontents = Column(LargeBinary)

    @exposed
    def fields(self, fields):
        fields.surname = Field(label='Surname', required=True)
        fields.name = Field(label='Name', required=True)
        fields.dob = DateField(label='Date of Birth', required=True)
        fields.uploaded_files = FileField(
            allow_multiple=False, max_size_bytes=4 * 1000 * 1000, max_files=1
        )

    def save(self):
        self.bsfilename = self.uploaded_files.filename
        self.bscontents = self.uploaded_files.contents
        Session.add(self)

    @exposed('save')
    def events(self, events):
        events.save = Event(label='Save', action=Action(self.save))


class GraphPage(CustomerPage):
    def __init__(self, view, main_bookmarks):
        super().__init__(view, main_bookmarks)

        fig1 = self.create_line_chart_figure()
        self.body.add_child(Chart(view, fig1, 'line'))

    def create_line_chart_figure(self):
        # rec = Session.query(Customer).first()
        rec = Session.query(Customer).filter(
            Customer.id == Session.query(func.max(Customer.id))
        )
        # import pdb;pdb.set_trace()
        wb = load_workbook(filename=BytesIO(rec[0].bscontents))
        ws = wb.active
        months = [c.value for c in ws['A'][1:]]
        income = [c.value for c in ws['B'][1:]]
        expences = [c.value for c in ws['C'][1:]]
        fig = pg.Figure()
        fig.add_trace(pg.Scatter(x=months, y=income, name=ws['B'][0].value))
        fig.add_trace(pg.Scatter(x=months, y=expences, name=ws['C'][0].value))
        fig.update_layout(
            title=f"{rec[0].name} {rec[0].surname} - Income and Expenditure",
            hovermode="x unified",
            xaxis_title='Months',
            yaxis_title='Amount',
        )
        return fig
